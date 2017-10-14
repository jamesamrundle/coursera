import openpyxl
import sys
import os
from openpyxl import utils
import pprint
from collections import OrderedDict
from operator import itemgetter

os.chdir('/home/jamesrundle/anaconda3/envs/invdata')




prodsWpars = []
thepars = []
inventory = []
invcount = []
searchdict = {} ##items in inventory that appear on parsheet
par4onhand= {} ##listed pars for what items appear on both inv and parsheet
whatsonhand= {}## full list of inventory: count for that item
actualonhand = {}##list of items in inventory that appear on par, and what we have of that item

def parproducts(): #generates list of items contained in par sheet and associated par(at same index value)
    wb = openpyxl.load_workbook('./Roof-par.xlsx')

    ws = wb.get_sheet_by_name('Par list')

    numofcols = ws.max_column
    numofrows = ws.max_row
    
    for i in range(2,numofrows+1):
        prodsWpars.append(ws['e'+str(i)].value)
        thepars.append(ws['a'+str(i)].value)

def inventoryproducts():#generates list of items contained in inventory sheet and associated count(at same index value)
    wb = openpyxl.load_workbook('./DR-inventory.xlsx')

    ws = wb.get_sheet_by_name('ROOFTOP KITCHEN')

    numofcols = ws.max_column
    numofrows = ws.max_row
    
    
    global inventory
    for i in range(6,123):
        inventory.append(ws['a'+str(i)].value)
        invcount.append(ws['f'+str(i)].value)
        whatsonhand[ws['a'+str(i)].value] = ws['f'+str(i)].value

inventoryproducts()

def propagatesearchstr(items): #uses par list(smaller list)
  for i in items:
    #print('#')
    stringbreak(i)

def stringbreak(product): #takes first word of par list
  searchstr = ''            
  global searchdict
  for i in product:
      if i not in (' ' , '-'):
                
                searchstr += i
                
      else:
                searchdict[product] = searchstr #propagates a list to check the inventory
                break

parproducts()
propagatesearchstr(prodsWpars)


def parforonhand():  ### uses searchdict to check for item in inventory
    ### searches for item based off first word from par list.
    for item in searchdict:
        for prod in inventory:
            if searchdict[item] in prod: ## if found makes dict of item:par and item:invcount
                par4onhand[prod] = thepars[prodsWpars.index(item)]
                actualonhand[prod] = invcount[inventory.index(prod)]
parforonhand()


def printdata():
    for i in par4onhand:
        print(i+ '--Par is: ' + str(par4onhand[i]) +' With :'+ str(actualonhand[i])+ ' --In current inventory')

printdata()        