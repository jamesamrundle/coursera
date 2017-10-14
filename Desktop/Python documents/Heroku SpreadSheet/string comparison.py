import openpyxl
import sys
import os
from openpyxl import utils
import pprint
from collections import OrderedDict
from operator import itemgetter

os.chdir('/home/jamesrundle/anaconda3/envs/invdata')




parprods = []
invprods = []
searchdict = {}
def parproducts():
    wb = openpyxl.load_workbook('./Roof-par.xlsx')

    ws = wb.get_sheet_by_name('Par list')

    numofcols = ws.max_column
    numofrows = ws.max_row
    
    for i in range(2,numofrows+1):
        parprods.append(ws['e'+str(i)].value)

def propagatesearchstr(items):
  for i in items:
    stringbreak(i)

def stringbreak(product):
  searchstr = ''
  for i in product:
      if i not in (' ' ,'-'):
                searchstr += i
      else:
                searchdict[product] = searchstr

propagatesearchstr(parprods)
