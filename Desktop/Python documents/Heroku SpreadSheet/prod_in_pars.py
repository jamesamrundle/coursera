import openpyxl
import sys
import os
from openpyxl import utils
import pprint
from collections import OrderedDict
from operator import itemgetter

os.chdir('/home/jamesrundle/anaconda3/envs/invdata')




parprods = []
invprods = []

def parproducts():
    wb = openpyxl.load_workbook('./Roof-par.xlsx')

    ws = wb.get_sheet_by_name('Par list')

    numofcols = ws.max_column
    numofrows = ws.max_row
    
    
    global parprods
    for i in range(2,numofrows+1):
        parprods.append(ws['e'+str(i)].value)
    
parproducts()        

def inventoryproducts():
    wb = openpyxl.load_workbook('./DR-inventory.xlsx')

    ws = wb.get_sheet_by_name('ROOFTOP KITCHEN')

    numofcols = ws.max_column
    numofrows = ws.max_row
    
    
    global invprods
    for i in range(6,123):
        invprods.append(ws['a'+str(i)].value)

inventoryproducts()
