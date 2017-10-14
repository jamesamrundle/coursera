#! /usr/bin/python3
# reads Censusexcel.py and tabulates population data for each county

import openpyxl
import pprint


wb = openpyxl.load_workbook('/home/jamesrundle/Desktop/Python documents/automate_online-materials/censuspopdata.xlsx')

print('Opening workbook \'Census Data\'.')
states = {}
countyData = {}
poptotal = 0

ws =  wb.active


for row in range(2,ws.max_row+1):
        
        state= ws['b'+ str(row)].value
        county = ws['c' + str(row)].value
        population = ws['d'+str(row)].value

        countyData[county] = 0
        states[state] = []
for row in range(2,ws.max_row+1):
        
        state= ws['b'+ str(row)].value
        county = ws['c' + str(row)].value
        population = ws['d'+str(row)].value

        countyData[county] += population
        #print(states)

        if [county] not in states[state]: ##adds each county
                states[state].append([county])
        
#print(countyData)
for i in states:
        print( i , len(states[i]))
##ACCURATELY GETS POP PER COUNTY
